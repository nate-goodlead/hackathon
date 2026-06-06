#!/usr/bin/env python3
"""
Parse real hackathon xlsx/zip data from data/incoming/ into pipeline CSV files.

Location mapping (not in source files — provided by data owner):
  portfolio company data      → Heeze   (Exact)
  portfolio company 2 data    → Brunssum (Yuki / Peter Ummels)
  Altis dataset 1.xlsx        → Andijk  (Gilde)
  Altis dataset 2.xlsx        → Winschoten (Exact)
"""

from __future__ import annotations

import csv
import json
import re
import zipfile
from calendar import monthrange
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

from fetch_weather import fetch_all_weather

ROOT = Path(__file__).resolve().parent.parent
INCOMING = ROOT / "data" / "incoming"
EXTRACTED = INCOMING / "extracted"
RAW = ROOT / "data" / "raw"
OUT = ROOT / "data" / "output"

MONTHS_NL = {"Jan": 1, "Feb": 2, "Mrt": 3, "Apr": 4, "Mei": 5, "Jun": 6,
             "Jul": 7, "Aug": 8, "Sep": 9, "Okt": 10, "Nov": 11, "Dec": 12}

# GL → driver category (extend as accounts discovered)
GL_CATEGORY = {
    "4000": "materials", "4010": "materials", "4020": "materials",
    "5000": "subcontractors", "5010": "subcontractors",
    "8000": "billing", "8001": "billing", "8002": "billing",
    "8004": "billing", "8005": "billing",
    "80000": "billing", "80020": "billing",
    "9000": "overhead", "9010": "overhead",
}


def ensure_extracted() -> None:
    EXTRACTED.mkdir(parents=True, exist_ok=True)
    for z in INCOMING.glob("*.zip"):
        with zipfile.ZipFile(z) as archive:
            archive.extractall(EXTRACTED)


def load_locations() -> list[dict]:
    path = INCOMING / "opco_locations.json"
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return []


def iso(d: date | datetime) -> str:
    if isinstance(d, datetime):
        return d.date().isoformat()
    return d.isoformat()


def gl_category(gl: str) -> str:
    gl = str(gl).strip()
    return GL_CATEGORY.get(gl, "billing" if gl.startswith("8") else "unmapped")


def write_gl_mapping(discovered: set[str]) -> None:
    RAW.mkdir(parents=True, exist_ok=True)
    rows = []
    for gl in sorted(discovered, key=lambda x: (len(x), x)):
        cat = gl_category(gl)
        rows.append([gl, cat, f"Auto-mapped from real data — {cat}"])
    with (RAW / "gl_account_mapping.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["gl_account", "category", "description"])
        w.writerows(rows)


def parse_heeze() -> list[dict]:
    folder = EXTRACTED / "portfolio company data"
    if not folder.exists():
        return []
    rows = []
    opco = "Portfolio Company Heeze"
    city = "Heeze"
    for path in sorted(folder.glob("*.xlsx")):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            gl = str(int(row[0])) if isinstance(row[0], float) else str(row[0]).strip()
            txn_date = row[2]
            if not isinstance(txn_date, datetime):
                continue
            debet = float(row[5] or 0)
            credit = float(row[6] or 0)
            amount = round(credit - debet, 2)
            if amount == 0:
                continue
            desc = str(row[7] or "GB export")
            rows.append({
                "date": iso(txn_date),
                "gl_account": gl,
                "amount": amount,
                "description": desc,
                "opco": opco,
                "project_id": f"PRJ-{city.upper()[:3]}-001",
                "source_system": "Exact",
                "city": city,
            })
        wb.close()
    return rows


def parse_brunssum() -> list[dict]:
    folder = EXTRACTED / "portfolio company 2 data"
    if not folder.exists():
        return []
    rows = []
    opco = "Dakdekkersbedrijf Peter Ummels"
    city = "Brunssum"
    for path in sorted(folder.glob("*.xlsx")):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        gl_code = None
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if row and row[0] == "Grootboekrekening" and row[1]:
                gl_code = str(row[1]).split(" - ")[0].strip()
            if row and row[0] == "Nr.":
                header_row = i
                break
        if not gl_code or header_row is None:
            wb.close()
            continue
        for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
            if not row or not isinstance(row[2], datetime):
                continue
            debet = float(row[5] or 0)
            credit = float(row[6] or 0)
            amount = round(credit - debet, 2)
            if amount == 0:
                continue
            dagboek = str(row[4] or "")
            rows.append({
                "date": iso(row[2]),
                "gl_account": gl_code,
                "amount": amount,
                "description": dagboek,
                "opco": opco,
                "project_id": f"PRJ-{city.upper()[:3]}-001",
                "source_system": "Yuki",
                "city": city,
            })
        wb.close()
    return rows


def parse_andijk() -> list[dict]:
    path = EXTRACTED / "datasets" / "Altis dataset 1.xlsx"
    if not path.exists():
        return []
    rows = []
    opco = "Portfolio Company Andijk"
    city = "Andijk"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = next(ws.iter_rows(max_row=1, values_only=True), None)
        if not header:
            continue
        months = [h for h in header[1:13] if h in MONTHS_NL]
        if not months:
            continue
        year = 2026 if "2026" in sheet_name else int(re.search(r"\d{4}", sheet_name).group()) if re.search(r"\d{4}", sheet_name) else 2025
        for row in ws.iter_rows(min_row=2, max_row=40, values_only=True):
            if not row or not row[0] or not isinstance(row[0], str):
                continue
            label = row[0].strip()
            if not label[0].isdigit():
                continue
            gl = label.split()[0]
            for col_idx, month_label in enumerate(header[1:13], start=1):
                if month_label not in MONTHS_NL:
                    continue
                val = row[col_idx] if col_idx < len(row) else None
                if not val or not isinstance(val, (int, float)):
                    continue
                month = MONTHS_NL[month_label]
                day = min(15, monthrange(year, month)[1])
                txn_date = date(year, month, day)
                amount = round(float(val), 2)
                rows.append({
                    "date": iso(txn_date),
                    "gl_account": gl,
                    "amount": amount,
                    "description": label,
                    "opco": opco,
                    "project_id": f"PRJ-{city.upper()[:3]}-001",
                    "source_system": "Gilde",
                    "city": city,
                })
    wb.close()
    return rows


def parse_winschoten() -> list[dict]:
    path = EXTRACTED / "datasets" / "Altis dataset 2.xlsx"
    if not path.exists():
        return []
    rows = []
    opco = "Portfolio Company Winschoten"
    city = "Winschoten"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet_name in ("2024", "2025", "2026"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not isinstance(row[0], datetime):
                continue
            debet = float(row[3] or 0) if len(row) > 3 else 0
            credit = float(row[4] or 0) if len(row) > 4 else 0
            amount = round(credit - debet, 2)
            if amount == 0:
                continue
            dagboek = str(row[2] or "journal")
            gl = "8000" if "Verkoop" in dagboek else "9000"
            rows.append({
                "date": iso(row[0]),
                "gl_account": gl,
                "amount": amount,
                "description": dagboek,
                "opco": opco,
                "project_id": f"PRJ-{city.upper()[:4]}-001",
                "source_system": "Exact",
                "city": city,
            })
    wb.close()
    return rows


def build_wip_csv(locations: list[dict], unified: list[dict]) -> None:
    """Build WIP projects — one primary project per opco/location."""
    RAW.mkdir(parents=True, exist_ok=True)
    rows = []
    milestones = [
        ("Membrane install", 0.30, 85000),
        ("Insulation complete", 0.55, 120000),
        ("Handover", 1.00, 95000),
    ]
    billing_by_opco: dict[str, float] = {}
    for r in unified:
        if r.get("gl_category") == "billing" or str(r.get("gl_account", "")).startswith("8"):
            billing_by_opco[r["opco"]] = billing_by_opco.get(r["opco"], 0) + abs(r["amount"])

    for i, loc in enumerate(locations):
        opco = loc["opco_name"]
        city = loc["city"]
        contract = max(400_000, billing_by_opco.get(opco, 500_000) * 1.2)
        pct = 0.20 + (i * 0.08)
        wip = round(contract * pct)
        pid = f"PRJ-{city.upper()[:4]}-001"
        seg = "large" if contract > 600_000 else "small"
        start_w = 1 + (i % 3)
        for idx, (mname, mpct, mval) in enumerate(milestones):
            rows.append([
                pid, f"{city} Roofing Project", opco, round(contract), wip, round(pct * 100, 1),
                mname, mpct, mval, start_w + idx * 2, seg, city,
            ])
    with (RAW / "projects_wip.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "project_id", "project_name", "opco", "contract_value", "wip_to_date", "pct_complete",
            "milestone_name", "milestone_pct", "milestone_billing", "scheduled_week",
            "customer_segment", "city",
        ])
        w.writerows(rows)


def write_unified(rows: list[dict], locations: list[dict]) -> None:
    discovered = {r["gl_account"] for r in rows}
    write_gl_mapping(discovered)

    for r in rows:
        r["gl_category"] = gl_category(r["gl_account"])

    headers = ["date", "gl_account", "amount", "description", "opco", "project_id", "source_system", "gl_category", "city"]
    OUT.mkdir(parents=True, exist_ok=True)
    with (OUT / "unified_data.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)

    mapping_rows = [{"gl_account": gl, "category": gl_category(gl), "status": "mapped"} for gl in sorted(discovered)]
    unmapped = [gl for gl in discovered if gl_category(gl) == "unmapped"]
    for gl in unmapped:
        mapping_rows.append({"gl_account": gl, "category": "unmapped", "status": "flagged"})

    with (OUT / "gl_mapping.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["gl_account", "category", "status"])
        w.writeheader()
        w.writerows(mapping_rows)

    notes = [
        "Real data ingestion — Altis Groep Hackathon",
        "",
        "Sources parsed from data/incoming/ zip files:",
        "  portfolio company data     → Heeze (Exact) — location mapped by data owner",
        "  portfolio company 2 data   → Brunssum / Peter Ummels (Yuki)",
        "  Altis dataset 1.xlsx       → Andijk (Gilde monthly P&L)",
        "  Altis dataset 2.xlsx       → Winschoten (Exact journal)",
        "",
        f"Total unified rows: {len(rows)}",
        f"Opcos: {len({r['opco'] for r in rows})}",
        f"Cities: {', '.join(sorted({r['city'] for r in rows}))}",
        "",
        "Location column added — source files do not include company location.",
        "Weather.csv generated per city for schedule-delay modelling.",
    ]
    (OUT / "data_notes.txt").write_text("\n".join(notes), encoding="utf-8")

    public = ROOT / "public" / "data"
    public.mkdir(parents=True, exist_ok=True)
    for name in ("unified_data.csv", "gl_mapping.csv", "data_notes.txt"):
        (public / name).write_text((OUT / name).read_text(encoding="utf-8"), encoding="utf-8")
    (public / "opco_locations.json").write_text(
        json.dumps(locations, indent=2), encoding="utf-8"
    )

    print(f"Parsed {len(rows)} rows from real incoming data")


def main() -> None:
    ensure_extracted()
    locations = load_locations()

    all_rows = (
        parse_heeze()
        + parse_brunssum()
        + parse_andijk()
        + parse_winschoten()
    )
    if not all_rows:
        print("No xlsx data found in data/incoming/extracted/ — run with zip files present")
        return

    # Filter to recent 13-week window for forecast relevance
    cutoff = date.today() - timedelta(weeks=52)
    recent = [r for r in all_rows if date.fromisoformat(r["date"]) >= cutoff]
    rows = recent if len(recent) > 100 else all_rows

    write_unified(rows, locations)

    print("Fetching real weather from Open-Meteo…")
    try:
        fetch_all_weather(locations)
    except Exception as e:
        print(f"Weather fetch failed ({e}) — re-run: npm run data:weather")

    if not (RAW / "covenant_terms.json").exists():
        default_covenant = {
            "headroom_threshold_eur": 500_000,
            "interest_coverage_minimum": 2.0,
            "current_interest_coverage": 2.4,
            "net_debt_eur": 4_200_000,
            "ebitda_annual_eur": 1_800_000,
            "formula": "headroom = headroom_threshold_eur - projected_net_debt_increase",
        }
        (RAW / "covenant_terms.json").write_text(json.dumps(default_covenant, indent=2), encoding="utf-8")

    build_wip_csv(locations, rows)

    # Copy locations for frontend
    (INCOMING / "opco_locations.json").write_text(
        json.dumps(locations, indent=2), encoding="utf-8"
    )
    print(f"Cities: {', '.join(sorted({r['city'] for r in rows}))}")


if __name__ == "__main__":
    main()
