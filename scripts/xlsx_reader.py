"""Parse Excel uploads into tabular rows for the ingest pipeline."""

from __future__ import annotations

import csv
import io
import re
from calendar import monthrange
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from zipfile import BadZipFile

import openpyxl


@dataclass
class SheetData:
    name: str
    headers: list[str]
    rows: list[dict[str, str]]
    score: int = 0


FORECAST_HEADERS = [
    "date",
    "gl_account",
    "amount",
    "description",
    "opco",
    "project_id",
    "source_system",
    "city",
]

MONTHS_NL = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "mrt": 3,
    "mar": 3,
    "apr": 4,
    "mei": 5,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "sept": 9,
    "okt": 10,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


def _norm_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]", "", h.lower().strip())


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _clean_gl(value) -> str:
    text = _cell_str(value)
    if not text:
        return ""
    match = re.search(r"\d{3,6}", text)
    return match.group(0) if match else ""


def _is_number(value) -> bool:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
    try:
        float(str(value).replace(".", "").replace(",", "."))
        return True
    except (TypeError, ValueError):
        return False


def _month_from_value(value) -> int | None:
    text = _cell_str(value).lower().strip(". ")
    return MONTHS_NL.get(text)


def _year_from_sheet(sheet_name: str) -> int | None:
    match = re.search(r"(20\d{2})", sheet_name)
    return int(match.group(1)) if match else None


def _month_end(year: int, month: int) -> str:
    return date(year, month, monthrange(year, month)[1]).isoformat()


def _find_header(headers: list[str], candidates: list[str]) -> str | None:
    normalized = {_norm_header(h): h for h in headers}
    for candidate in candidates:
        norm_candidate = _norm_header(candidate)
        for norm_header, header in normalized.items():
            if norm_header == norm_candidate or norm_candidate in norm_header:
                return header
    return None


def _standard_row_dict(headers: list[str], values: list[str]) -> dict[str, str]:
    padded = values + [""] * max(0, len(headers) - len(values))
    return dict(zip(headers, padded[: len(headers)]))


def _try_monthly_pnl(ws, sheet_name: str) -> tuple[list[str], list[dict[str, str]], int] | None:
    """Wide P&L sheets: first row has Jan-Dec, rows contain GL labels."""
    year = _year_from_sheet(sheet_name)
    if not year:
        return None

    scanned = list(ws.iter_rows(values_only=True))
    header_idx: int | None = None
    month_cols: list[tuple[int, int]] = []
    for idx, row in enumerate(scanned[:8]):
        found = [(col_idx, month) for col_idx, value in enumerate(row) if (month := _month_from_value(value))]
        if len(found) >= 3:
            header_idx = idx
            month_cols = found
            break
    if header_idx is None:
        return None

    rows: list[dict[str, str]] = []
    for row in scanned[header_idx + 1 :]:
        label = _cell_str(row[0] if row else "")
        gl = _clean_gl(label)
        if not gl:
            continue
        for col_idx, month in month_cols:
            value = row[col_idx] if col_idx < len(row) else None
            if not _is_number(value) or float(str(value).replace(".", "").replace(",", ".")) == 0:
                continue
            rows.append(
                {
                    "date": _month_end(year, month),
                    "gl_account": gl,
                    "amount": _cell_str(value),
                    "description": f"{label} · {sheet_name}",
                    "opco": "",
                    "project_id": "",
                    "source_system": "",
                    "city": "",
                }
            )
    if not rows:
        return None
    return FORECAST_HEADERS, rows, len(rows) + 450


def _try_date_columns_summary(ws, sheet_name: str) -> tuple[list[str], list[dict[str, str]], int] | None:
    """Summary sheets: one header row of dates with one or more numeric rows."""
    scanned = list(ws.iter_rows(values_only=True))
    for header_idx, row in enumerate(scanned[:8]):
        date_cols: list[tuple[int, str]] = []
        for col_idx, value in enumerate(row):
            if isinstance(value, (datetime, date)):
                date_cols.append((col_idx, _cell_str(value)))
            elif re.fullmatch(r"\d{4}-\d{2}-\d{2}", _cell_str(value)):
                date_cols.append((col_idx, _cell_str(value)))
        if len(date_cols) < 3:
            continue

        rows: list[dict[str, str]] = []
        for value_row in scanned[header_idx + 1 : header_idx + 6]:
            numeric_hits = sum(
                1
                for col_idx, _ in date_cols
                if col_idx < len(value_row) and _is_number(value_row[col_idx])
            )
            if numeric_hits < 3:
                continue
            for col_idx, txn_date in date_cols:
                amount = value_row[col_idx] if col_idx < len(value_row) else None
                if not _is_number(amount) or float(str(amount).replace(".", "").replace(",", ".")) == 0:
                    continue
                rows.append(
                    {
                        "date": txn_date,
                        "gl_account": "8000",
                        "amount": _cell_str(amount),
                        "description": f"Monthly revenue summary · {sheet_name}",
                        "opco": "",
                        "project_id": "",
                        "source_system": "",
                        "city": "",
                    }
                )
            break
        if rows:
            return FORECAST_HEADERS, rows, len(rows) + 300
    return None


def _maybe_normalize_sales_journal(
    headers: list[str],
    rows: list[dict[str, str]],
    sheet_name: str,
) -> tuple[list[str], list[dict[str, str]], int] | None:
    """Sales journals without a GL account: keep credit-side billing rows only."""
    has_gl = bool(_find_header(headers, ["rekening", "grootboek", "gl_account", "account"]))
    date_col = _find_header(headers, ["datum", "date", "boekingsdatum"])
    credit_col = _find_header(headers, ["credit", "creditbedrag"])
    journal_col = _find_header(headers, ["dagboek"])
    desc_col = _find_header(headers, ["boekingstekst", "omschrijving", "bkst.nr.", "boeknummer", "dagboek"])
    doc_col = _find_header(headers, ["bkst.nr.", "boeknummer", "boekstuk"])
    if has_gl or not date_col or not credit_col or not journal_col:
        return None

    normalized: list[dict[str, str]] = []
    for row in rows:
        journal = row.get(journal_col, "").lower()
        credit = row.get(credit_col, "")
        if ("verkoop" not in journal and "omzet" not in journal) or not credit or not _is_number(credit):
            continue
        doc = row.get(doc_col, "") if doc_col else ""
        desc = row.get(desc_col, "") if desc_col else ""
        description = " · ".join(part for part in [desc or "Sales journal billing", doc, sheet_name] if part)
        normalized.append(
            {
                "date": row.get(date_col, ""),
                "gl_account": "8000",
                "amount": credit,
                "description": description,
                "opco": "",
                "project_id": "",
                "source_system": "",
                "city": "",
            }
        )
    if not normalized:
        return None
    return FORECAST_HEADERS, normalized, len(normalized) + 350


def _maybe_normalize_invoice_list(
    headers: list[str],
    rows: list[dict[str, str]],
    sheet_name: str,
) -> tuple[list[str], list[dict[str, str]], int] | None:
    """Invoice lists without GL: treat invoice amount as billing."""
    has_invoice_marker = any("factuur" in h.lower() or "invoice" in h.lower() for h in headers)
    if not has_invoice_marker:
        return None
    has_gl = bool(_find_header(headers, ["rekening", "grootboek", "gl_account", "account"]))
    date_col = _find_header(headers, ["factuurdatum", "invoice_date", "datum", "date"])
    amount_col = _find_header(headers, ["factuurbedrag", "invoice_amount", "bedrag", "amount"])
    invoice_col = _find_header(headers, ["factuurnummer", "invoice_number", "invoice"])
    if has_gl or not date_col or not amount_col:
        return None

    normalized: list[dict[str, str]] = []
    for row in rows:
        amount = row.get(amount_col, "")
        txn_date = row.get(date_col, "")
        if not txn_date or not amount or not _is_number(amount):
            continue
        invoice = row.get(invoice_col, "") if invoice_col else ""
        normalized.append(
            {
                "date": txn_date,
                "gl_account": "8000",
                "amount": amount,
                "description": f"Invoice billing {invoice} · {sheet_name}".strip(),
                "opco": "",
                "project_id": "",
                "source_system": "",
                "city": "",
            }
        )
    if not normalized:
        return None
    return FORECAST_HEADERS, normalized, len(normalized) + 325


def _try_yuki_fintransactions(ws) -> tuple[list[str], list[dict[str, str]], int] | None:
    """Yuki exports: metadata rows then header row starting with Nr."""
    gl_code = ""
    header_idx: int | None = None
    scanned: list[tuple] = []
    for row in ws.iter_rows(values_only=True):
        scanned.append(row)
        if row and row[0] == "Grootboekrekening" and row[1]:
            gl_code = str(row[1]).split(" - ")[0].strip()
        if header_idx is None and row and str(row[0]).strip() == "Nr.":
            header_idx = len(scanned) - 1
    if header_idx is None:
        return None

    header_row = scanned[header_idx]
    headers = [_cell_str(h) or f"col_{j}" for j, h in enumerate(header_row)]
    if "gl_account" not in headers:
        headers = [*headers, "gl_account"]

    rows: list[dict[str, str]] = []
    for row in scanned[header_idx + 1 :]:
        values = [_cell_str(v) for v in row]
        if not any(values):
            continue
        if values[0] == "" and not values[2]:
            continue
        padded = values + [""] * max(0, len(headers) - 1 - len(values))
        record = dict(zip(headers[:-1], padded[: len(headers) - 1]))
        record["gl_account"] = gl_code
        rows.append(record)
    return headers, rows, len(rows) + 500


def _read_standard_sheet(ws, max_rows: int | None = None) -> tuple[list[str], list[dict[str, str]], int] | None:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return None
    headers = [_cell_str(h) or f"col_{i}" for i, h in enumerate(header_row)]
    if not any(headers):
        return None

    rows: list[dict[str, str]] = []
    for i, row in enumerate(rows_iter):
        if max_rows is not None and i >= max_rows:
            break
        values = [_cell_str(v) for v in row]
        if not any(values):
            continue
        rows.append(_standard_row_dict(headers, values))

    invoice_list = _maybe_normalize_invoice_list(headers, rows, ws.title)
    if invoice_list:
        return invoice_list

    sales_journal = _maybe_normalize_sales_journal(headers, rows, ws.title)
    if sales_journal:
        return sales_journal

    score = len(rows) + (
        10
        if any(
            "grootboek" in h.lower() or "account" in h.lower() or h.lower() == "rekening"
            for h in headers
        )
        else 0
    )
    return headers, rows, score


def _read_worksheet(ws, sheet_name: str, max_rows: int | None = None) -> SheetData | None:
    yuki = _try_yuki_fintransactions(ws)
    if yuki:
        headers, rows, score = yuki
        if max_rows is not None:
            rows = rows[:max_rows]
        if not rows:
            return None
        return SheetData(sheet_name, headers, rows, score)

    monthly_pnl = _try_monthly_pnl(ws, sheet_name)
    if monthly_pnl:
        headers, rows, score = monthly_pnl
        if max_rows is not None:
            rows = rows[:max_rows]
        if not rows:
            return None
        return SheetData(sheet_name, headers, rows, score)

    date_summary = _try_date_columns_summary(ws, sheet_name)
    if date_summary:
        headers, rows, score = date_summary
        if max_rows is not None:
            rows = rows[:max_rows]
        if not rows:
            return None
        return SheetData(sheet_name, headers, rows, score)

    standard = _read_standard_sheet(ws, max_rows=max_rows)
    if not standard:
        return None
    headers, rows, score = standard
    if not rows:
        return None
    return SheetData(sheet_name, headers, rows, score)


def xlsx_to_all_sheets(content: bytes, max_rows: int | None = None) -> list[SheetData]:
    """Read every worksheet that contains tabular data."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except BadZipFile as exc:
        raise ValueError("Invalid Excel workbook — upload a saved .xlsx file, not an Excel lock/temp file.") from exc
    sheets: list[SheetData] = []
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            sheet = _read_worksheet(ws, name, max_rows=max_rows)
            if sheet:
                sheets.append(sheet)
    finally:
        wb.close()
    if any(not s.name.lower().strip().startswith("totaal") for s in sheets):
        sheets = [s for s in sheets if not s.name.lower().strip().startswith("totaal")]
    return sheets


def _union_headers(sheets: list[SheetData]) -> list[str]:
    primary = max(sheets, key=lambda s: len(s.rows))
    union = list(primary.headers)
    for sheet in sheets:
        for h in sheet.headers:
            if h not in union and h != "_source_sheet":
                union.append(h)
    return union


def combine_sheets(
    sheets: list[SheetData],
) -> tuple[list[str], list[dict[str, str]], list[dict], list[str]]:
    """Merge all usable sheets into one row set with per-sheet metadata."""
    usable = [s for s in sheets if s.rows]
    if not usable:
        raise ValueError("No usable worksheet found in Excel file")

    multi = len(usable) > 1
    headers = _union_headers(usable)
    if multi and "_source_sheet" not in headers:
        headers.append("_source_sheet")

    combined: list[dict[str, str]] = []
    breakdown: list[dict] = []
    for sheet in usable:
        for row in sheet.rows:
            merged = {h: row.get(h, "") for h in headers if h != "_source_sheet"}
            if multi:
                merged["_source_sheet"] = sheet.name
            combined.append(merged)
        breakdown.append(
            {
                "sheetName": sheet.name,
                "rowCount": len(sheet.rows),
                "headers": sheet.headers,
            }
        )

    sheet_names = [s.name for s in usable]
    return headers, combined, breakdown, sheet_names


def xlsx_to_combined_rows(
    content: bytes, max_rows: int | None = None
) -> tuple[list[str], list[dict[str, str]], list[dict], list[str]]:
    """Return merged headers/rows from every usable worksheet."""
    sheets = xlsx_to_all_sheets(content, max_rows=max_rows)
    return combine_sheets(sheets)


def xlsx_to_rows(content: bytes, max_rows: int | None = None) -> tuple[list[str], list[dict[str, str]], str]:
    """Return headers, rows, and a sheet label (all sheets combined)."""
    headers, rows, breakdown, sheet_names = xlsx_to_combined_rows(content, max_rows=max_rows)
    if len(sheet_names) == 1:
        label = sheet_names[0]
    else:
        label = ", ".join(sheet_names)
    return headers, rows, label


def rows_to_csv_bytes(headers: list[str], rows: list[dict[str, str]]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def save_xlsx_as_csv(
    content: bytes, dest: Path
) -> tuple[list[str], list[dict[str, str]], str | None, list[str], list[dict]]:
    """Convert all worksheets to one CSV; returns combined data + per-sheet breakdown."""
    headers, rows, breakdown, sheet_names = xlsx_to_combined_rows(content)
    dest.write_bytes(rows_to_csv_bytes(headers, rows))
    label = sheet_names[0] if len(sheet_names) == 1 else ", ".join(sheet_names)
    return headers, rows, label, sheet_names, breakdown
